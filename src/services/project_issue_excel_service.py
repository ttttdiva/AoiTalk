"""Project issue Excel reader used by project-management agent tools."""

from __future__ import annotations

import hashlib
import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional


HEADER_ALIASES = {
    "number": {"no", "no.", "番号", "#"},
    "kind": {"区分", "分類", "type", "種別"},
    "phase": {"フェーズ", "phase", "工程"},
    "status": {"status", "ステータス", "状態", "状況"},
    "created_at": {"起票日", "作成日", "created", "created at"},
    "reporter": {"起票者", "報告者", "reporter"},
    "importance": {"重要度", "priority", "importance"},
    "title": {"課題概要", "概要", "件名", "タイトル", "title", "summary"},
    "detail": {"課題詳細", "詳細", "内容", "detail", "description"},
    "action_plan": {"actionplan", "action plan", "対応方針", "対策", "アクション"},
    "close_condition": {"課題close条件", "close condition", "完了条件"},
    "due_date": {"対応期限", "期限", "期日", "due", "due date"},
    "owner": {"主担当者", "担当者", "対応者", "owner", "assignee"},
    "history": {
        "対応経緯",
        "経緯",
        "回答・対応",
        "回答・対応（更新は赤字追記)",
        "history",
        "対応履歴",
    },
    "resolved_at": {"対策終了日", "対応完了日", "完了日", "resolved", "resolved at"},
    "approved_at": {"完了承認日", "承認日", "approved", "approved at"},
    "approver": {"完了承認者", "承認者", "approver"},
    "notes": {"備考", "notes", "note"},
}


@dataclass
class IssueTableRow:
    source_key: str
    row_hash: str
    file_path: str
    sheet_name: str
    row_number: int
    number: Optional[int]
    kind: Optional[str]
    phase: Optional[str]
    status: str
    created_at: Optional[str]
    reporter: Optional[str]
    importance: Optional[str]
    title: str
    detail: Optional[str]
    action_plan: Optional[str]
    close_condition: Optional[str]
    due_date: Optional[str]
    owner: Optional[str]
    history: Optional[str]
    resolved_at: Optional[str]
    approved_at: Optional[str]
    approver: Optional[str]
    notes: Optional[str]
    raw: Optional[dict[str, Any]] = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _to_iso_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20_000 <= float(value) <= 60_000:
        return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return None
    y, m, d = match.groups()
    return f"{y}-{m.zfill(2)}-{d.zfill(2)}"


def _row_hash(raw: dict[str, Any]) -> str:
    payload = json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def _project_root() -> Path:
    return Path(os.environ.get("AOITALK_PROJECT_ROOT", ".")).resolve()


def _workspaces_root() -> Path:
    configured = os.environ.get("AOITALK_WORKSPACES_DIR")
    return Path(configured).resolve() if configured else _project_root() / "workspaces"


def _project_storage_root(project_context: dict[str, Any]) -> Optional[Path]:
    workspace_root = _clean(project_context.get("workspace_root"))
    if workspace_root:
        return Path(workspace_root).resolve()
    project_id = _clean(project_context.get("id"))
    if not project_id:
        return None
    return (_workspaces_root() / "_projects" / f"project_{project_id}").resolve()


def _normalize_project_file_path(value: Any, project_id: Optional[str] = None) -> Optional[str]:
    text = _clean(value)
    if not text:
        return None
    text = text.replace("\\", "/").rstrip("/")
    if project_id:
        prefix = f"_projects/project_{project_id}/"
        if text.startswith(prefix):
            text = text[len(prefix):]
        elif text.startswith("_projects/project_"):
            return None
    elif text.startswith("_projects/project_"):
        match = re.match(r"^_projects/project_[^/]+/(.*)$", text, flags=re.IGNORECASE)
        if match:
            text = match.group(1)
    if re.match(r"^[A-Za-z]:/", text) or text.startswith("/") or text.startswith("//"):
        return None
    parts = [part for part in text.split("/") if part not in {"", "."}]
    if not parts or any(part == ".." for part in parts):
        return None
    return "/".join(parts)


def _issue_candidate_score(path: Path, storage_root: Path) -> tuple[int, float, int]:
    relative = str(path.relative_to(storage_root)).replace("\\", "/")
    haystack = f"{path.name}\n{relative}".casefold()
    score = 0
    if "課題管理" in haystack:
        score += 50
    if "課題" in haystack:
        score += 30
    if "issue" in haystack:
        score += 20
    location_score = 1 if "management/" in relative.casefold() else 0
    return (score, path.stat().st_mtime, location_score)


def resolve_issue_path(
    project_context: dict[str, Any],
    issue_file: str = "",
) -> tuple[Optional[Path], Optional[str]]:
    project_id = _clean(project_context.get("id"))
    storage_root = _project_storage_root(project_context)
    if storage_root is None:
        return None, None

    explicit = _normalize_project_file_path(issue_file, project_id)
    if explicit:
        resolved = (storage_root / explicit).resolve()
        try:
            resolved.relative_to(storage_root)
        except ValueError:
            return None, explicit
        return (resolved if resolved.exists() else None), explicit

    candidates: list[Path] = []
    configured = _normalize_project_file_path(project_context.get("issue_file"), project_id)
    if configured:
        configured_path = (storage_root / configured).resolve()
        if configured_path.exists():
            candidates.append(configured_path)

    if storage_root.exists():
        for path in storage_root.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
                continue
            haystack = f"{path.name}\n{path.relative_to(storage_root)}".casefold()
            if "課題" in haystack or "issue" in haystack:
                candidates.append(path.resolve())

    unique = {str(path): path for path in candidates}
    if not unique:
        return None, configured
    selected = max(unique.values(), key=lambda item: _issue_candidate_score(item, storage_root))
    return selected, str(selected.relative_to(storage_root)).replace("\\", "/")


def _detect_header(rows: list[list[Any]]) -> Optional[tuple[int, dict[str, int]]]:
    best: Optional[tuple[int, int, dict[str, int]]] = None
    for row_index, row in enumerate(rows[:30]):
        columns: dict[str, int] = {}
        score = 0
        for field, aliases in HEADER_ALIASES.items():
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            for col_index, cell in enumerate(row):
                if _normalize_header(cell) in normalized_aliases:
                    columns[field] = col_index
                    score += 5 if field in {"number", "title", "status"} else 1
                    break
        if best is None or score > best[1]:
            best = (row_index, score, columns)
    if not best or best[1] < 8 or "title" not in best[2]:
        return None
    return best[0], best[2]


def _cell(row: list[Any], columns: dict[str, int], field: str) -> Any:
    index = columns.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_number(value: Any) -> Optional[int]:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _clean(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def _title_from_row(row: list[Any], columns: dict[str, int]) -> Optional[str]:
    title = _clean(_cell(row, columns, "title"))
    if title:
        return title
    detail = _clean(_cell(row, columns, "detail"))
    if detail:
        return detail.splitlines()[0][:200]
    return None


def read_issue_rows(
    project_context: dict[str, Any],
    issue_file: str = "",
) -> tuple[list[IssueTableRow], list[str], Optional[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on local environment
        return [], [f"openpyxl を読み込めません: {exc}"], None

    file_path, relative_file_path = resolve_issue_path(project_context, issue_file)
    if file_path is None:
        target = relative_file_path or issue_file or project_context.get("issue_file") or ""
        suffix = f": {target}" if target else ""
        return [], [f"課題管理表ファイルが見つかりません{suffix}"], relative_file_path

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    rows: list[IssueTableRow] = []

    for sheet in workbook.worksheets:
        numbered_rows: list[tuple[int, list[Any]]] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if any(_clean(value) for value in values):
                numbered_rows.append((row_number, values))
        values = [row for _, row in numbered_rows]
        detected = _detect_header(values)
        if not detected:
            continue
        header_index, columns = detected
        for compact_index in range(header_index + 1, len(numbered_rows)):
            row_number, row = numbered_rows[compact_index]
            title = _title_from_row(row, columns)
            number = _parse_number(_cell(row, columns, "number"))
            if not title or number is None:
                continue
            raw = {field: _cell(row, columns, field) for field in columns}
            source_key = f"{relative_file_path}::{sheet.title}::{number}"
            resolved_at = _to_iso_date(_cell(row, columns, "resolved_at"))
            status = _clean(_cell(row, columns, "status"))
            if not status:
                status = "完了" if resolved_at else "未着手"
            rows.append(
                IssueTableRow(
                    source_key=source_key,
                    row_hash=_row_hash(raw),
                    file_path=relative_file_path or file_path.name,
                    sheet_name=sheet.title,
                    row_number=row_number,
                    number=number,
                    kind=_clean(_cell(row, columns, "kind")),
                    phase=_clean(_cell(row, columns, "phase")),
                    status=status,
                    created_at=_to_iso_date(_cell(row, columns, "created_at")),
                    reporter=_clean(_cell(row, columns, "reporter")),
                    importance=_clean(_cell(row, columns, "importance")),
                    title=title,
                    detail=_clean(_cell(row, columns, "detail")),
                    action_plan=_clean(_cell(row, columns, "action_plan")),
                    close_condition=_clean(_cell(row, columns, "close_condition")),
                    due_date=_to_iso_date(_cell(row, columns, "due_date")),
                    owner=_clean(_cell(row, columns, "owner")),
                    history=_clean(_cell(row, columns, "history")),
                    resolved_at=resolved_at,
                    approved_at=_to_iso_date(_cell(row, columns, "approved_at")),
                    approver=_clean(_cell(row, columns, "approver")),
                    notes=_clean(_cell(row, columns, "notes")),
                    raw=raw,
                )
            )

    if not rows:
        return [], ["課題管理表として読めるシートまたは課題行が見つかりませんでした"], relative_file_path
    return rows, [], relative_file_path


def is_closed_issue(row: IssueTableRow) -> bool:
    return row.status.casefold() in {"完了", "closed", "done", "close"}


def summarize_issue_rows(rows: list[IssueTableRow]) -> dict[str, Any]:
    status_counts: dict[str, int] = {}
    importance_counts: dict[str, int] = {}
    for row in rows:
        status_counts[row.status] = status_counts.get(row.status, 0) + 1
        if row.importance:
            importance_counts[row.importance] = importance_counts.get(row.importance, 0) + 1
    open_rows = [row for row in rows if not is_closed_issue(row)]
    return {
        "total": len(rows),
        "open_count": len(open_rows),
        "closed_count": len(rows) - len(open_rows),
        "status_counts": status_counts,
        "importance_counts": importance_counts,
    }
