"""WBS Excel reader used by project-management agent tools."""

from __future__ import annotations

import hashlib
import json
import os
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional


HEADER_ALIASES = {
    "wbs_id": {"wbs", "wbs番号", "id", "no", "番号", "#"},
    "title": {"タスク名", "作業項目", "作業内容", "項目", "件名", "task", "name", "title"},
    "description": {"説明", "備考", "内容", "詳細", "description", "note"},
    "assignee": {"担当", "担当者", "担当部署", "assignee", "owner"},
    "status": {"状態", "ステータス", "status", "状況"},
    "planned_start": {"予定開始日", "開始予定", "開始日", "計画開始", "planned start", "start"},
    "planned_end": {"予定終了日", "終了予定", "終了日", "期限", "期日", "due", "planned end", "end"},
    "actual_start": {"実績開始日", "実開始", "actual start"},
    "actual_end": {"実績終了日", "実終了", "完了日", "actual end"},
    "progress": {"進捗率", "進捗", "progress", "%"},
    "request_text": {"確認事項", "要確認", "依頼事項", "顧客確認", "確認先", "qa"},
}


@dataclass
class WbsTaskRow:
    source_key: str
    row_hash: str
    file_path: str
    sheet_name: str
    row_number: int
    wbs_id: Optional[str]
    title: str
    description: Optional[str]
    assignee: Optional[str]
    status: str
    priority: str
    planned_start: Optional[str]
    planned_end: Optional[str]
    actual_start: Optional[str]
    actual_end: Optional[str]
    progress: Optional[float]
    request_text: Optional[str]
    raw: Optional[dict[str, Any]] = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class LayeredWbsHeader:
    header_index: int
    first_data_index: int
    title_columns: list[int]
    assignee_columns: list[dict[str, Any]]
    planned_start: Optional[int]
    planned_end: Optional[int]
    actual_start: Optional[int]
    actual_end: Optional[int]
    progress: Optional[int]


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _to_iso_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20_000 <= float(value) <= 60_000:
        # Excel serial dates use 1899-12-30 as the practical epoch.
        return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return None
    y, m, d = match.groups()
    return f"{y}-{m.zfill(2)}-{d.zfill(2)}"


def _parse_progress(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) / 100 if value > 1 else float(value)
    match = re.search(r"(\d+(?:\.\d+)?)", str(value))
    if not match:
        return None
    numeric = float(match.group(1))
    return numeric / 100 if "%" in str(value) or numeric > 1 else numeric


def _map_status(value: Any, progress: Optional[float], actual_end: Optional[str]) -> str:
    text = str(value or "").lower()
    if actual_end or (progress is not None and progress >= 1):
        return "closed"
    if re.search(r"確認|review|承認|回答待ち", text):
        return "review"
    if re.search(r"保留|hold|中断|待ち", text):
        return "on_hold"
    if re.search(r"進行|対応中|着手|doing|progress", text):
        return "in_progress"
    return "open"


def _map_priority(planned_end: Optional[str], status: str) -> str:
    if not planned_end or status == "closed":
        return "medium"
    try:
        due = datetime.fromisoformat(planned_end).date()
    except ValueError:
        return "medium"
    days = (due - date.today()).days
    if days < 0:
        return "urgent"
    if days <= 3:
        return "high"
    return "medium"


def _detect_header(rows: list[list[Any]]) -> Optional[tuple[int, dict[str, int]]]:
    best: Optional[tuple[int, int, dict[str, int]]] = None
    for row_index, row in enumerate(rows[:20]):
        columns: dict[str, int] = {}
        score = 0
        for field, aliases in HEADER_ALIASES.items():
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            for col_index, cell in enumerate(row):
                if _normalize_header(cell) in normalized_aliases:
                    columns[field] = col_index
                    score += 3 if field == "title" else 1
                    break
        if best is None or score > best[1]:
            best = (row_index, score, columns)
    if not best or best[1] < 3 or "title" not in best[2]:
        return None
    return best[0], best[2]


def _find_header_column(
    row: list[Any],
    aliases: set[str] | tuple[str, ...] | list[str],
    start: int = 0,
    end: Optional[int] = None,
) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    stop = len(row) if end is None else min(end, len(row))
    for index in range(start, stop):
        if _normalize_header(row[index]) in normalized_aliases:
            return index
    return None


def _min_defined(*values: Optional[int]) -> Optional[int]:
    candidates = [value for value in values if value is not None]
    return min(candidates) if candidates else None


def _detect_layered_wbs_header(rows: list[list[Any]]) -> Optional[LayeredWbsHeader]:
    """Detect Gantt-style WBS sheets with hierarchy columns and grouped date headers."""
    for row_index, row in enumerate(rows[:30]):
        next_row = rows[row_index + 1] if row_index + 1 < len(rows) else []
        wbs_index = _find_header_column(row, {"wbs", "wbs番号"})
        if wbs_index is None:
            continue

        assignee_start = _find_header_column(row, HEADER_ALIASES["assignee"])
        planned_start_group = _find_header_column(row, {"予定", "計画"})
        actual_start_group = _find_header_column(row, {"実績"})
        progress = _find_header_column(row, HEADER_ALIASES["progress"])
        if progress is None:
            progress = _find_header_column(next_row, HEADER_ALIASES["progress"])
        planned_end_boundary = (
            _min_defined(actual_start_group, progress, len(row)) or len(row)
        )
        actual_end_boundary = _min_defined(progress, len(row)) or len(row)

        planned_start = (
            None
            if planned_start_group is None
            else _find_header_column(
                next_row,
                {"開始日", "開始予定", "planned start", "start"},
                planned_start_group,
                planned_end_boundary,
            )
        )
        planned_end = (
            None
            if planned_start_group is None
            else _find_header_column(
                next_row,
                {
                    "終了日",
                    "終了予定",
                    "予定終了日",
                    "期限",
                    "due",
                    "planned end",
                    "end",
                },
                planned_start_group,
                planned_end_boundary,
            )
        )
        actual_start = (
            None
            if actual_start_group is None
            else _find_header_column(
                next_row,
                {"開始日", "実績開始日", "actual start"},
                actual_start_group,
                actual_end_boundary,
            )
        )
        actual_end = (
            None
            if actual_start_group is None
            else _find_header_column(
                next_row,
                {"終了日", "完了日", "実績終了日", "actual end"},
                actual_start_group,
                actual_end_boundary,
            )
        )
        if planned_start is None and planned_end is None and actual_start is None:
            continue

        title_end = _min_defined(
            assignee_start,
            planned_start_group,
            actual_start_group,
            progress,
        ) or len(row)
        title_columns = list(range(wbs_index, max(wbs_index, title_end)))
        if not title_columns:
            continue

        assignee_end = _min_defined(planned_start_group, actual_start_group, progress) or len(row)
        assignee_columns = []
        if assignee_start is not None:
            for index in range(assignee_start, max(assignee_start, assignee_end)):
                assignee_columns.append(
                    {
                        "index": index,
                        "label": _clean(
                            next_row[index] if index < len(next_row) else None
                        )
                        or (
                            None
                            if index == assignee_start
                            else _clean(row[index] if index < len(row) else None)
                        ),
                    }
                )

        return LayeredWbsHeader(
            header_index=row_index,
            first_data_index=row_index + 2,
            title_columns=title_columns,
            assignee_columns=assignee_columns,
            planned_start=planned_start,
            planned_end=planned_end,
            actual_start=actual_start,
            actual_end=actual_end,
            progress=progress,
        )

    return None


def _cell(row: list[Any], columns: dict[str, int], field: str) -> Any:
    index = columns.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _is_wbs_number_part(value: str) -> bool:
    return bool(re.match(r"^\d+(?:\.\d+)?$", value.strip()))


def _format_wbs_number_part(value: str) -> str:
    trimmed = value.strip()
    return str(int(float(trimmed))) if re.match(r"^\d+\.0+$", trimmed) else trimmed


def _extract_layered_assignee(
    row: list[Any],
    columns: list[dict[str, Any]],
) -> Optional[str]:
    assignees: list[str] = []
    for column in columns:
        index = int(column["index"])
        marker = _clean(row[index] if index < len(row) else None)
        if not marker:
            continue
        label = column.get("label")
        if re.match(r"^[●○〇◎◯✓✔■]+$", marker):
            if label:
                assignees.append(str(label))
        else:
            assignees.append(f"{label}: {marker}" if label else marker)
    if not assignees:
        return None
    return ", ".join(dict.fromkeys(assignees))


def _row_hash(data: dict[str, Any]) -> str:
    payload = json.dumps(data, ensure_ascii=False, default=str, sort_keys=True)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def _non_empty_sheet_rows(sheet: Any) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(row)
        if any(_clean(value) for value in values):
            rows.append((row_number, values))
    return rows


def _read_layered_wbs_rows(
    numbered_rows: list[tuple[int, list[Any]]],
    header: LayeredWbsHeader,
    relative_file_path: str,
    sheet_name: str,
) -> list[WbsTaskRow]:
    rows: list[WbsTaskRow] = []
    hierarchy: dict[int, str] = {}

    for compact_index in range(header.first_data_index, len(numbered_rows)):
        row_number, row = numbered_rows[compact_index]
        for column in header.title_columns:
            value = _clean(row[column] if column < len(row) else None)
            if not value:
                continue
            hierarchy[column] = value
            for deeper_column in header.title_columns:
                if deeper_column > column:
                    hierarchy.pop(deeper_column, None)

        hierarchy_values = [
            hierarchy[column]
            for column in header.title_columns
            if hierarchy.get(column)
        ]
        title = next(
            (
                value
                for value in reversed(hierarchy_values)
                if not _is_wbs_number_part(value)
            ),
            None,
        )
        if not title:
            continue

        planned_start = _to_iso_date(
            row[header.planned_start]
            if header.planned_start is not None and header.planned_start < len(row)
            else None
        )
        planned_end = _to_iso_date(
            row[header.planned_end]
            if header.planned_end is not None and header.planned_end < len(row)
            else None
        )
        actual_start = _to_iso_date(
            row[header.actual_start]
            if header.actual_start is not None and header.actual_start < len(row)
            else None
        )
        actual_end = _to_iso_date(
            row[header.actual_end]
            if header.actual_end is not None and header.actual_end < len(row)
            else None
        )
        progress = _parse_progress(
            row[header.progress]
            if header.progress is not None and header.progress < len(row)
            else None
        )
        assignee = _extract_layered_assignee(row, header.assignee_columns)
        if (
            not planned_start
            and not planned_end
            and not actual_start
            and not actual_end
            and progress is None
            and not assignee
        ):
            continue

        wbs_id_parts = [
            _format_wbs_number_part(value)
            for value in hierarchy_values
            if _is_wbs_number_part(value)
        ]
        wbs_id = ".".join(wbs_id_parts) if wbs_id_parts else None
        description_parts = [
            value
            for value in hierarchy_values
            if not _is_wbs_number_part(value) and value != title
        ]
        raw = {
            "hierarchy": hierarchy_values,
            "assignee": assignee,
            "planned_start": (
                row[header.planned_start]
                if header.planned_start is not None and header.planned_start < len(row)
                else None
            ),
            "planned_end": (
                row[header.planned_end]
                if header.planned_end is not None and header.planned_end < len(row)
                else None
            ),
            "actual_start": (
                row[header.actual_start]
                if header.actual_start is not None and header.actual_start < len(row)
                else None
            ),
            "actual_end": (
                row[header.actual_end]
                if header.actual_end is not None and header.actual_end < len(row)
                else None
            ),
            "progress": (
                row[header.progress]
                if header.progress is not None and header.progress < len(row)
                else None
            ),
        }
        status = _map_status(None, progress, actual_end)
        source_key = f"{relative_file_path}::{sheet_name}::{wbs_id or f'row-{row_number}'}"
        rows.append(
            WbsTaskRow(
                source_key=source_key,
                row_hash=_row_hash(raw),
                file_path=relative_file_path,
                sheet_name=sheet_name,
                row_number=row_number,
                wbs_id=wbs_id,
                title=title,
                description=" > ".join(description_parts) if description_parts else None,
                assignee=assignee,
                status=status,
                priority=_map_priority(planned_end, status),
                planned_start=planned_start,
                planned_end=planned_end,
                actual_start=actual_start,
                actual_end=actual_end,
                progress=progress,
                request_text=None,
                raw=raw,
            )
        )

    return rows


def _project_root() -> Path:
    return Path(os.environ.get("AOITALK_PROJECT_ROOT", ".")).resolve()


def _workspaces_root() -> Path:
    configured = os.environ.get("AOITALK_WORKSPACES_DIR")
    return Path(configured).resolve() if configured else _project_root() / "workspaces"


def _project_storage_root(project_context: dict[str, Any]) -> Optional[Path]:
    workspace_root = _clean(project_context.get("workspace_root"))
    if workspace_root:
        return Path(workspace_root).resolve()
    project_id = _clean(project_context.get("id"))
    if not project_id:
        return None
    return (_workspaces_root() / "_projects" / f"project_{project_id}").resolve()


def _normalize_project_file_path(value: Any, project_id: Optional[str] = None) -> Optional[str]:
    text = _clean(value)
    if not text:
        return None
    text = text.replace("\\", "/").rstrip("/")
    if project_id:
        prefix = f"_projects/project_{project_id}/"
        if text.startswith(prefix):
            text = text[len(prefix):]
        elif text.startswith("_projects/project_"):
            return None
    elif text.startswith("_projects/project_"):
        match = re.match(r"^_projects/project_[^/]+/(.*)$", text, flags=re.IGNORECASE)
        if match:
            text = match.group(1)
    if re.match(r"^[A-Za-z]:/", text) or text.startswith("/") or text.startswith("//"):
        return None
    parts = [part for part in text.split("/") if part not in {"", "."}]
    if not parts or any(part == ".." for part in parts):
        return None
    return "/".join(parts)


def resolve_wbs_path(project_context: dict[str, Any]) -> Optional[Path]:
    project_id = _clean(project_context.get("id"))
    storage_root = _project_storage_root(project_context)
    wbs_file = project_context.get("wbs_file")
    relative_wbs_file = _normalize_project_file_path(wbs_file, project_id)
    if not relative_wbs_file or storage_root is None:
        return None
    resolved = (storage_root / relative_wbs_file).resolve()
    try:
        resolved.relative_to(storage_root)
    except ValueError:
        return None
    return resolved


def read_wbs_file(
    file_path: Path,
    *,
    relative_file_path: Optional[str] = None,
) -> tuple[list[WbsTaskRow], list[str]]:
    """Read every task row from a supported WBS workbook."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on local environment
        return [], [f"openpyxl を読み込めません: {exc}"]

    path = Path(file_path).expanduser().resolve()
    source_path = relative_file_path or path.name
    if not path.exists():
        return [], [f"WBSファイルが見つかりません: {source_path}"]
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return [], [f"未対応のWBS形式です: {path.suffix or '(拡張子なし)'}"]

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[WbsTaskRow] = []

    try:
        for sheet in workbook.worksheets:
            numbered_rows = _non_empty_sheet_rows(sheet)
            values = [row for _, row in numbered_rows]
            detected = _detect_header(values)
            if not detected:
                layered = _detect_layered_wbs_header(values)
                if layered:
                    rows.extend(
                        _read_layered_wbs_rows(
                            numbered_rows,
                            layered,
                            source_path,
                            sheet.title,
                        )
                    )
                continue
            header_index, columns = detected
            for compact_index in range(header_index + 1, len(numbered_rows)):
                index, row = numbered_rows[compact_index]
                title = _clean(_cell(row, columns, "title"))
                if not title:
                    continue
                planned_start = _to_iso_date(_cell(row, columns, "planned_start"))
                planned_end = _to_iso_date(_cell(row, columns, "planned_end"))
                actual_start = _to_iso_date(_cell(row, columns, "actual_start"))
                actual_end = _to_iso_date(_cell(row, columns, "actual_end"))
                progress = _parse_progress(_cell(row, columns, "progress"))
                status = _map_status(_cell(row, columns, "status"), progress, actual_end)
                raw = {field: _cell(row, columns, field) for field in columns}
                wbs_id = _clean(_cell(row, columns, "wbs_id"))
                source_key = f"{source_path}::{sheet.title}::{wbs_id or f'row-{index}'}"
                rows.append(
                    WbsTaskRow(
                        source_key=source_key,
                        row_hash=_row_hash(raw),
                        file_path=source_path,
                        sheet_name=sheet.title,
                        row_number=index,
                        wbs_id=wbs_id,
                        title=title,
                        description=_clean(_cell(row, columns, "description")),
                        assignee=_clean(_cell(row, columns, "assignee")),
                        status=status,
                        priority=_map_priority(planned_end, status),
                        planned_start=planned_start,
                        planned_end=planned_end,
                        actual_start=actual_start,
                        actual_end=actual_end,
                        progress=progress,
                        request_text=_clean(_cell(row, columns, "request_text")),
                        raw=raw,
                    )
                )
    finally:
        workbook.close()

    if not rows:
        return [], ["WBSとして読めるシートまたはタスク行が見つかりませんでした"]
    return rows, []


def read_wbs_rows(project_context: dict[str, Any]) -> tuple[list[WbsTaskRow], list[str]]:
    file_path = resolve_wbs_path(project_context)
    if file_path is None:
        return [], ["WBSファイルが設定されていません"]
    relative_file_path = _normalize_project_file_path(
        project_context.get("wbs_file"),
        _clean(project_context.get("id")),
    ) or file_path.name
    return read_wbs_file(file_path, relative_file_path=relative_file_path)


def summarize_request_items(rows: list[WbsTaskRow], limit: int = 20) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    pattern = re.compile(r"確認|依頼|回答待ち|未回答|QA|Q&A|問い合わせ", re.IGNORECASE)
    for row in rows:
        haystack = " ".join(
            item for item in [row.title, row.description, row.request_text, row.status] if item
        )
        if not pattern.search(haystack):
            continue
        target = "customer" if re.search(r"顧客|お客様|客先|customer|client", haystack, re.I) else "unknown"
        items.append(
            {
                "title": row.request_text or row.title,
                "target": target,
                "reason": row.title if row.request_text else "WBS上で確認または依頼が必要な状態です",
                "source_path": row.file_path,
                "source_ref": f"{row.sheet_name}!{row.row_number}",
                "due_at": row.planned_end,
                "status": "blocked" if row.status == "on_hold" else "draft",
            }
        )
        if len(items) >= limit:
            break
    return items
