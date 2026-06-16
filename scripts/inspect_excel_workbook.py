"""Inspect an Excel workbook and print a bounded JSON structure summary."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def inspect_workbook(
    workbook_path: Path,
    *,
    max_rows: int = 10,
    max_columns: int = 30,
    data_only: bool = False,
) -> dict[str, Any]:
    path = workbook_path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"未対応のExcel形式です: {path.suffix or '(拡張子なし)'}。"
            " .xlsx/.xlsm/.xltx/.xltm を指定してください。"
        )
    if max_rows < 0 or max_columns < 1:
        raise ValueError("max_rows は0以上、max_columns は1以上で指定してください。")

    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(path, data_only=data_only, keep_vba=keep_vba)
    vba_archive = getattr(workbook, "vba_archive", None)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            sample_rows: list[dict[str, Any]] = []
            if max_rows:
                for row_index, row in enumerate(
                    sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row, max_rows),
                        max_col=min(sheet.max_column, max_columns),
                        values_only=True,
                    ),
                    start=1,
                ):
                    sample_rows.append(
                        {
                            "row": row_index,
                            "values": [_json_value(value) for value in row],
                        }
                    )
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "merged_cells": [str(item) for item in sheet.merged_cells.ranges],
                    "sample_rows": sample_rows,
                }
            )
        return {
            "path": str(path),
            "file_type": path.suffix.lower(),
            "keep_vba": keep_vba,
            "data_only": data_only,
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()
        if vba_archive is not None:
            vba_archive.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Excelブックのシート構造と先頭行をJSONで表示します。"
    )
    parser.add_argument("workbook", type=Path, help=".xlsx/.xlsmファイル")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=10,
        help="各シートから出力する先頭行数。0で行データを省略します。",
    )
    parser.add_argument(
        "--max-columns",
        type=int,
        default=30,
        help="各行から出力する最大列数。",
    )
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="数式そのものではなく、保存済みの計算結果を読みます。",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        result = inspect_workbook(
            args.workbook,
            max_rows=args.max_rows,
            max_columns=args.max_columns,
            data_only=args.data_only,
        )
    except (FileNotFoundError, ValueError, OSError) as exc:
        print(json.dumps({"success": False, "error": str(exc)}, ensure_ascii=False))
        return 1

    print(
        json.dumps(
            {"success": True, **result},
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
