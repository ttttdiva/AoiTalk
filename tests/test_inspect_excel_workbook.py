from datetime import date
import importlib.util
from pathlib import Path

import pytest
from openpyxl import Workbook

SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "inspect_excel_workbook.py"
SPEC = importlib.util.spec_from_file_location("inspect_excel_workbook", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)
inspect_workbook = MODULE.inspect_workbook


def test_inspect_workbook_returns_sheet_structure_and_sample_rows(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "パラメータ"
    sheet.append(["項目", "値", "確認日"])
    sheet.append(["OS", "Windows Server 2025", date(2026, 6, 14)])
    sheet.merge_cells("A4:B4")

    path = tmp_path / "parameters.xlsx"
    workbook.save(path)
    workbook.close()

    result = inspect_workbook(path, max_rows=2, max_columns=3)

    assert result["sheet_count"] == 1
    assert result["keep_vba"] is False
    assert result["sheets"][0]["name"] == "パラメータ"
    assert result["sheets"][0]["merged_cells"] == ["A4:B4"]
    assert result["sheets"][0]["sample_rows"][1]["values"] == [
        "OS",
        "Windows Server 2025",
        "2026-06-14T00:00:00",
    ]


def test_inspect_workbook_keeps_vba_for_xlsm(tmp_path: Path):
    workbook = Workbook()
    workbook.active.append(["WBS", "タスク"])
    path = tmp_path / "wbs.xlsm"
    workbook.save(path)
    workbook.close()

    result = inspect_workbook(path, max_rows=1)

    assert result["file_type"] == ".xlsm"
    assert result["keep_vba"] is True


def test_inspect_workbook_rejects_legacy_xls(tmp_path: Path):
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not-an-xls-file")

    with pytest.raises(ValueError, match="未対応のExcel形式"):
        inspect_workbook(path)
