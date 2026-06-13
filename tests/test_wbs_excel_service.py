from datetime import date
from pathlib import Path

from openpyxl import Workbook

from src.services.wbs_excel_service import read_wbs_rows, summarize_request_items


def test_read_wbs_rows_detects_headers_and_status(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WBS"
    sheet.append(["WBS番号", "タスク名", "予定終了日", "進捗率", "状態", "担当", "確認事項"])
    sheet.append(["1.1", "設計書レビュー", "2026-05-01", "50%", "確認待ち", "田中", "顧客に承認可否を確認"])
    sheet.append(["1.2", "実装", "2026-05-10", "100%", "完了", "佐藤", None])

    wbs_path = tmp_path / "WBS.xlsx"
    workbook.save(wbs_path)

    rows, errors = read_wbs_rows(
        {
            "workspace_root": str(tmp_path),
            "wbs_file": "WBS.xlsx",
        }
    )

    assert errors == []
    assert len(rows) == 2
    assert rows[0].wbs_id == "1.1"
    assert rows[0].status == "review"
    assert rows[0].planned_end == "2026-05-01"
    assert rows[0].progress == 0.5
    assert rows[1].status == "closed"


def test_read_wbs_rows_detects_layered_wbs_headers(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "進捗管理"
    sheet.append(["案件名"])
    sheet.append([])
    sheet.append(
        [
            "WBS",
            None,
            None,
            None,
            None,
            "担当",
            "予定",
            None,
            "実績",
            None,
            "進捗率",
        ]
    )
    sheet.append(
        [
            None,
            None,
            None,
            None,
            None,
            "NOS",
            "開始日",
            "終了日",
            "開始日",
            "終了日",
            None,
        ]
    )
    sheet.append([1, "プロジェクト計画"])
    sheet.append([None, None, 1, "キックオフ"])
    sheet.append(
        [
            None,
            None,
            None,
            None,
            "キックオフ資料作成",
            "●",
            date(2026, 4, 13),
            date(2026, 4, 17),
            date(2026, 4, 13),
            date(2026, 4, 16),
            1,
        ]
    )

    wbs_path = tmp_path / "WBS.xlsm"
    workbook.save(wbs_path)

    rows, errors = read_wbs_rows(
        {
            "workspace_root": str(tmp_path),
            "wbs_file": "WBS.xlsm",
        }
    )

    assert errors == []
    assert len(rows) == 1
    row = rows[0]
    assert row.title == "キックオフ資料作成"
    assert row.description == "プロジェクト計画 > キックオフ"
    assert row.assignee == "NOS"
    assert row.planned_end == "2026-04-17"
    assert row.actual_end == "2026-04-16"
    assert row.status == "closed"


def test_summarize_request_items_from_wbs_rows(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["WBS番号", "タスク名", "確認事項"])
    sheet.append(["2.1", "VPN接続先確認", "お客様に接続先IPを確認"])

    wbs_path = tmp_path / "WBS.xlsx"
    workbook.save(wbs_path)

    rows, _ = read_wbs_rows({"workspace_root": str(tmp_path), "wbs_file": "WBS.xlsx"})
    items = summarize_request_items(rows)

    assert len(items) == 1
    assert items[0]["target"] == "customer"
    assert items[0]["title"] == "お客様に接続先IPを確認"


def test_read_wbs_rows_rejects_paths_outside_workspace(tmp_path: Path):
    outside = tmp_path.parent / "outside-wbs.xlsx"
    workbook = Workbook()
    workbook.active.append(["WBS番号", "タスク名"])
    workbook.active.append(["9.1", "外部ファイル"])
    workbook.save(outside)

    rows, errors = read_wbs_rows(
        {
            "workspace_root": str(tmp_path),
            "wbs_file": "../outside-wbs.xlsx",
        }
    )

    assert rows == []
    assert errors == ["WBSファイルが設定されていません"]
